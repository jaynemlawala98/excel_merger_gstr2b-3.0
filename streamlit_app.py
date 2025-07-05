# streamlit_app.py

import streamlit as st
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
import io

st.set_page_config(page_title="Excel Merger GSTR2B", layout="centered")

st.title("Excel Merger - GSTR2B")
st.markdown("Created by **Jay Nemlawala**")

uploaded_files = st.file_uploader(
    "Upload multiple Excel files (GSTR2B format)",
    type=["xlsx"],
    accept_multiple_files=True
)

sheet_configs = [
    ("B2B", 6),
    ("B2BA", 7),
    ("B2B-CDNR", 6),
    ("B2B-CDNRA", 7)
]

if uploaded_files and st.button("Merge Files"):
    output = io.BytesIO()
    output_wb = openpyxl.Workbook()

    for sheet_name, skip_rows in sheet_configs:
        output_ws = output_wb.active if sheet_name == "B2B" else output_wb.create_sheet(sheet_name)
        first = True

        for uploaded_file in uploaded_files:
            wb = openpyxl.load_workbook(uploaded_file)
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]

            if first:
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=skip_rows), start=1):
                    for col_idx, cell in enumerate(row, start=1):
                        new_cell = output_ws.cell(row=row_idx, column=col_idx, value=cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)
                for merged in ws.merged_cells.ranges:
                    output_ws.merge_cells(str(merged))
                first = False

            for row in ws.iter_rows(min_row=skip_rows + 1):
                output_ws.append([cell.value for cell in row])

        # Auto column width
        for col in output_ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            output_ws.column_dimensions[col_letter].width = max_length + 5

    output_wb.save(output)
    output.seek(0)

    st.success("✅ Files merged successfully!")

    st.download_button(
        label="Download Merged Excel File",
        data=output,
        file_name="merged_gstr2b.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
