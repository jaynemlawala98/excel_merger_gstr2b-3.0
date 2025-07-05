import streamlit as st
import openpyxl
import os
import io
import zipfile
from openpyxl.utils import get_column_letter
from copy import copy
import tempfile

# Set page config
st.set_page_config(
    page_title="Excel Merger GSTR2B",
    page_icon="📊",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# Custom CSS for better styling
st.markdown("""
<style>
    .main > div {
        padding-top: 2rem;
    }
    
    .stTitle {
        text-align: center;
        color: #1f77b4;
        font-size: 2.5rem;
        margin-bottom: 2rem;
    }
    
    .file-list {
        background-color: #ffffff;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
        min-height: 200px;
        border: 2px solid #1f77b4;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    
    .file-list h4 {
        color: #1f77b4;
        margin-bottom: 0.5rem;
    }
    
    .file-list ul {
        list-style-type: none;
        padding-left: 0;
    }
    
    .file-list li {
        padding: 0.5rem;
        margin: 0.25rem 0;
        background-color: #f8f9fa;
        border-radius: 0.25rem;
        border-left: 4px solid #1f77b4;
    }
    
    .footer {
        position: fixed;
        bottom: 10px;
        right: 20px;
        color: #666;
        font-size: 0.8rem;
    }
    
    .stButton > button {
        width: 100%;
        margin: 0.2rem 0;
    }
    
    .success-message {
        background-color: #d4edda;
        color: #155724;
        padding: 1rem;
        border-radius: 0.5rem;
        border: 1px solid #c3e6cb;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

# Initialize session state
if 'selected_files' not in st.session_state:
    st.session_state.selected_files = []
if 'file_contents' not in st.session_state:
    st.session_state.file_contents = {}

def update_file_list():
    """Update the file list display"""
    if st.session_state.selected_files:
        file_list_html = "<div class='file-list'><h4>Selected Files:</h4><ul>"
        for i, file_name in enumerate(st.session_state.selected_files, 1):
            file_list_html += f"<li>{i}. {file_name}</li>"
        file_list_html += "</ul></div>"
        st.markdown(file_list_html, unsafe_allow_html=True)
    else:
        st.markdown("<div class='file-list'><h4>Selected Files:</h4><p style='color: #666; font-style: italic;'>Use 'Select Files' button above to add Excel files.</p></div>", unsafe_allow_html=True)

def move_up():
    """Move the last file to the first position"""
    if len(st.session_state.selected_files) > 1:
        files = st.session_state.selected_files
        contents = st.session_state.file_contents
        
        # Move file name
        last_file = files.pop()
        files.insert(0, last_file)
        
        # Move file content
        last_content = contents[last_file]
        new_contents = {last_file: last_content}
        for file_name in files[1:]:
            new_contents[file_name] = contents[file_name]
        
        st.session_state.file_contents = new_contents

def move_down():
    """Move the first file to the last position"""
    if len(st.session_state.selected_files) > 1:
        files = st.session_state.selected_files
        contents = st.session_state.file_contents
        
        # Move file name
        first_file = files.pop(0)
        files.append(first_file)
        
        # Move file content
        first_content = contents[first_file]
        new_contents = {}
        for file_name in files[:-1]:
            new_contents[file_name] = contents[file_name]
        new_contents[first_file] = first_content
        
        st.session_state.file_contents = new_contents

def remove_file():
    """Remove the last file from the list"""
    if st.session_state.selected_files:
        removed_file = st.session_state.selected_files.pop()
        if removed_file in st.session_state.file_contents:
            del st.session_state.file_contents[removed_file]
        st.rerun()

def clear_all_files():
    """Clear all files from the list"""
    st.session_state.selected_files = []
    st.session_state.file_contents = {}
    st.rerun()

def merge_files():
    """Merge the selected Excel files"""
    if not st.session_state.selected_files:
        st.error("No files selected!")
        return None
    
    try:
        # Create progress bar
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        output_wb = openpyxl.Workbook()
        
        # Process each sheet type
        sheet_configs = [("B2B", 6), ("B2BA", 7), ("B2B-CDNR", 6), ("B2B-CDNRA", 7)]
        
        for sheet_idx, (sheet_name, skip_rows) in enumerate(sheet_configs):
            status_text.text(f"Processing {sheet_name} sheet...")
            progress_bar.progress((sheet_idx + 1) / len(sheet_configs))
            
            output_ws = output_wb.active if sheet_name == "B2B" else output_wb.create_sheet(sheet_name)
            
            first = True
            for file_name in st.session_state.selected_files:
                file_content = st.session_state.file_contents[file_name]
                
                # Load workbook from bytes
                wb = openpyxl.load_workbook(io.BytesIO(file_content))
                
                if sheet_name not in wb.sheetnames:
                    continue
                    
                ws = wb[sheet_name]
                
                # Copy headers and formatting from first file
                if first:
                    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=skip_rows), start=1):
                        for col_idx, cell in enumerate(row, start=1):
                            new_cell = output_ws.cell(row=row_idx, column=col_idx, value=cell.value)
                            if cell.has_style:
                                new_cell.font = copy(cell.font)
                                new_cell.border = copy(cell.border)
                                new_cell.fill = copy(cell.fill)
                                new_cell.number_format = copy(cell.number_format)
                                new_cell.protection = copy(cell.protection)
                                new_cell.alignment = copy(cell.alignment)
                    
                    # Copy merged cells
                    for merged in ws.merged_cells.ranges:
                        output_ws.merge_cells(str(merged))
                    
                    first = False
                
                # Copy data rows
                for row in ws.iter_rows(min_row=skip_rows+1):
                    output_ws.append([cell.value for cell in row])
            
            # Auto-adjust column widths
            for col in output_ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                output_ws.column_dimensions[col_letter].width = max_length + 5
        
        # Save to bytes
        output_buffer = io.BytesIO()
        output_wb.save(output_buffer)
        output_buffer.seek(0)
        
        progress_bar.progress(1.0)
        status_text.text("Files merged successfully!")
        
        return output_buffer.getvalue()
        
    except Exception as e:
        st.error(f"Error merging files: {str(e)}")
        return None

# Main App Layout
def main():
    # Title
    st.markdown("<h1 class='stTitle'>Excel Merger GSTR2B</h1>", unsafe_allow_html=True)
    
    # File upload section
    uploaded_files = st.file_uploader(
        "Select Excel Files",
        type=['xlsx'],
        accept_multiple_files=True,
        help="Select multiple Excel files to merge"
    )
    
    # Process uploaded files
    if uploaded_files:
        for uploaded_file in uploaded_files:
            if uploaded_file.name not in st.session_state.selected_files:
                st.session_state.selected_files.append(uploaded_file.name)
                st.session_state.file_contents[uploaded_file.name] = uploaded_file.read()
    
    # Display file list
    update_file_list()
    
    # Control buttons
    if st.session_state.selected_files:
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            if st.button("Move Up", help="Move last file to first position"):
                move_up()
                st.rerun()
        
        with col2:
            if st.button("Move Down", help="Move first file to last position"):
                move_down()
                st.rerun()
        
        with col3:
            if st.button("Remove File", help="Remove last file from list"):
                remove_file()
        
        with col4:
            if st.button("Clear All", help="Remove all files"):
                clear_all_files()
    
    # Merge button
    if st.session_state.selected_files:
        st.markdown("---")
        
        if st.button("🔄 Merge Files", type="primary", help="Merge all selected files"):
            with st.spinner("Merging files..."):
                merged_file = merge_files()
                
                if merged_file:
                    st.markdown("<div class='success-message'>✅ Files merged successfully!</div>", unsafe_allow_html=True)
                    
                    # Download button
                    st.download_button(
                        label="📥 Download Merged File",
                        data=merged_file,
                        file_name="merged_GSTR2B.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
    
    # Instructions
    with st.expander("📋 Instructions"):
        st.markdown("""
        1. **Select Files**: Upload multiple Excel files (.xlsx) that you want to merge
        2. **Arrange Files**: Use Move Up/Down buttons to change the order of files
        3. **Remove Files**: Use Remove File button to remove the last file or Clear All to remove all files
        4. **Merge**: Click 'Merge Files' to combine all selected files
        5. **Download**: Download the merged file when processing is complete
        
        **Supported Sheets**: B2B, B2BA, B2B-CDNR, B2B-CDNRA
        
        **Note**: The app will merge data from all files while preserving formatting from the first file.
        """)
    
    # Footer
    st.markdown("<div class='footer'>Created By Jay Nemlawala</div>", unsafe_allow_html=True)

if __name__ == "__main__":
    main()
